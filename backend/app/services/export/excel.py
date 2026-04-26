"""Export/import prices as Excel files using openpyxl."""

from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

HEADERS = ["Territory Code", "Territory Name", "Currency", "Customer Price", "Proceeds"]


class ExcelExportService:
    """Export/import prices as Excel files."""

    @staticmethod
    def export_prices(
        subscription_name: str,
        prices: list[dict],
    ) -> bytes:
        """Create an Excel workbook with price data.

        Args:
            subscription_name: Name used for the sheet title.
            prices: List of dicts with keys: territory_code, territory_name,
                    currency_code, customer_price, proceeds.

        Returns:
            The workbook serialised as bytes.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = subscription_name[:31]  # Excel sheet name limit

        # Write header row
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT

        # Sort prices by territory code for consistent output
        sorted_prices = sorted(prices, key=lambda p: p.get("territory_code", ""))

        for row_idx, price in enumerate(sorted_prices, start=2):
            ws.cell(row=row_idx, column=1, value=price.get("territory_code", ""))
            ws.cell(row=row_idx, column=2, value=price.get("territory_name", ""))
            ws.cell(row=row_idx, column=3, value=price.get("currency_code", ""))
            ws.cell(row=row_idx, column=4, value=price.get("customer_price", 0))
            ws.cell(row=row_idx, column=5, value=price.get("proceeds", 0))

        # Auto-size columns
        for col_idx in range(1, len(HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(HEADERS[col_idx - 1])
            for row in ws.iter_rows(
                min_row=2,
                max_row=ws.max_row,
                min_col=col_idx,
                max_col=col_idx,
            ):
                for cell in row:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 4

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def import_prices(file_bytes: bytes) -> list[dict]:
        """Parse an Excel file and return a list of price dicts.

        Expected format matches the export layout:
        Territory Code | Territory Name | Currency | Customer Price | Proceeds

        Returns:
            List of ``{territory_code: str, customer_price: float}``.
        """
        buf = io.BytesIO(file_bytes)
        wb = load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        result: list[dict] = []

        for row in rows:
            if not row or row[0] is None:
                continue
            territory_code = str(row[0]).strip()
            try:
                customer_price = float(row[3]) if row[3] is not None else 0.0
            except (ValueError, TypeError, IndexError):
                customer_price = 0.0

            result.append({
                "territory_code": territory_code,
                "customer_price": customer_price,
            })

        wb.close()
        return result
